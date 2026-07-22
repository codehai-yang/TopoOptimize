# -*- coding: utf-8 -*-
"""
线束拓扑闭环分析可视化工具
根据 BS4EM项目json优化设置.txt 的 edges 坐标 + 最终方案的 serviceableStatue，
可视化展示闭环位置。

用法：python visualize_cycle.py <txt路径> <方案json路径>
"""
import json
import os
import sys
import networkx as nx
import matplotlib.pyplot as plt


def load_edges(txt_path):
    """从 BS4EM项目json优化设置.txt 解析 edges
    文件可能末尾被污染（PowerShell 错误日志混入），用 raw_decode 只取第一个完整 JSON
    """
    with open(txt_path, 'r', encoding='utf-8') as f:
        content = f.read()
    # raw_decode 解析第一个完整 JSON 对象，自动跳过尾部污染
    decoder = json.JSONDecoder()
    obj, end_pos = decoder.raw_decode(content)
    print(f"  [load_edges] 解析到 {end_pos}/{len(content)} 字节 ({end_pos*100//len(content)}%)")
    edges = obj.get('edges', [])
    return edges, obj  # 顺便返回完整 obj，方便后续取 normList 等


def load_scheme(scheme_path, idx=0):
    """从方案 json 读 serviceableStatue（顺序对应 normList）
    兼容两种格式：
      1) 根是 dict: {"serviceableStatue": [...]}
      2) 根是 list: [{"serviceableStatue": [...]}, ...]  ← 取第 idx 个（默认 0）
    """
    with open(scheme_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if isinstance(data, list):
        if len(data) == 0:
            raise ValueError("方案文件为空 list")
        if idx >= len(data):
            raise ValueError(f"方案 idx={idx} 越界，list 长度={len(data)}")
        first = data[idx]
        if isinstance(first, dict):
            return first.get('serviceableStatue', first.get('statues', first.get('topologyStatusCodes', [])))
        else:
            raise ValueError(f"list 根第 {idx} 个元素不是 dict: {type(first)}")
    elif isinstance(data, dict):
        return data.get('serviceableStatue', data.get('statues', data.get('topologyStatusCodes', [])))
    else:
        raise ValueError(f"未识别的方案格式: type={type(data)}")


def load_scheme_from_excel(xlsx_path, scheme_idx, stage='绕线后'):
    """从方案状态变更追踪 Excel 读指定方案某阶段的状态列表。
    Excel 列布局(每方案 3 列 + 2 列间隔,1-based):
        方案 i: 精确前=(i-1)*5+1, 精确后=(i-1)*5+2, 绕线后=(i-1)*5+3
    注:Java 端写入时"分支 ID"列被后续状态写入覆盖(0-based 列 0 重复),
    汇总行 label 也被 k=0 覆盖为 "-",所以靠内容判断边界不可靠。
    用 max_row 反推:最后 3 行是总成本/总重量/总长度,数据行 = 6 .. max_row-3-1
    (中间还可能有一空行)。行序 = normList 顺序。
    """
    from openpyxl import load_workbook

    stage_to_col = {'精确前': 1, '精确后': 2, '绕线后': 3}
    if stage not in stage_to_col:
        raise ValueError(f"stage 必须是 '精确前'/'精确后'/'绕线后' 之一, 收到: {stage}")
    col = (scheme_idx - 1) * 5 + stage_to_col[stage]  # 1-based 列号

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = wb.active

    # 倒数 3 行是总成本/总重量/总长度,前 1 行是空行,数据只读到 max_row - 4
    data_end_row = max(6, sheet.max_row - 3)

    statuses = []
    for row in sheet.iter_rows(min_row=6, max_row=data_end_row, max_col=col, values_only=True):
        first_col = row[0] if row else None
        if first_col is None:
            continue
        first_str = str(first_col).strip()
        if not first_str:
            continue
        status = row[col - 1] if col <= len(row) else None
        statuses.append('-' if status is None else str(status).strip())
    wb.close()
    return statuses


def build_graph_and_find_cycles(edges, statuses):
    """
    1. 构造 networkx 图
    2. 标记 B 边（打断）
    3. 识别连通分量；>1 个 = 有闭环
    4. 找具体回路
    """
    G = nx.Graph()
    pos = {}

    for edge, status in zip(edges, statuses):
        sp = edge.get('startPointName', edge.get('startPoint'))
        ep = edge.get('endPointName', edge.get('endPoint'))
        if not sp or not ep:
            continue
        G.add_edge(sp, ep, status=status, edge_id=edge.get('id', ''))
        if sp not in pos:
            pos[sp] = (edge.get('startXCoordinate', 0), edge.get('startYCoordinate', 0))
        if ep not in pos:
            pos[ep] = (edge.get('endXCoordinate', 0), edge.get('endYCoordinate', 0))

    # B 边（打断）和 S 边（单线/绕线）都不参与闭环检测，只保留 C 边（与 Java recognizeLoopNew 对齐）
    break_edges = [(u, v) for u, v, d in G.edges(data=True) if d['status'] == 'B']
    skip_edges = [(u, v) for u, v, d in G.edges(data=True) if d['status'] in ('B', 'S')]

    # 移除 B 边后识别连通分量
    H = G.copy()
    H.remove_edges_from(break_edges)
    components = list(nx.connected_components(H))

    # 找回路：每个 B 边在原始图中找 u→v 路径，再闭合到 u
    # 同时跳过 B 边，避免回路高亮打到 B 边上（视觉混淆）
    cycles = []
    for u, v in break_edges:
        if u not in G or v not in G:
            continue
        # 构造临时图：去除 B 边
        H_cycle = G.copy()
        H_cycle.remove_edges_from(break_edges)
        # u 与 v 应在 H_cycle 中不连通（说明有回路）
        if u in H_cycle and v in H_cycle and nx.has_path(H_cycle, u, v) and not _in_same_component(H_cycle, u, v):
            try:
                # 找 u → v 的路径（只走 C/S 边）
                path = nx.shortest_path(H_cycle, u, v)
                # 闭合：u → ... → v → u（最后一个 B 边）
                closed = path + [u]
                cycles.append(closed)
            except nx.NetworkXNoPath:
                pass
        # 兜底：如果上面没找到，用原始 G 找 u→v 路径
        if not cycles or all(u not in c for c in cycles):
            try:
                path = nx.shortest_path(G, u, v)
                closed = path + [u]
                cycles.append(closed)
            except nx.NetworkXNoPath:
                pass

    # 去重：相同的节点序列只保留一个
    seen = set()
    unique_cycles = []
    for c in cycles:
        key = tuple(sorted(c[:-1]))  # 去掉末尾的 u，按无序集合去重
        if key not in seen:
            seen.add(key)
            unique_cycles.append(c)
    cycles = unique_cycles

    # 找所有基本环（与 Java 端 recognizeLoopNew 算法等价：DFS生成树 + 非树边 + LCA）
    # nx.cycle_basis 在数学上等价于 recognizeLoopNew，返回环基的独立环集合
    # 排除 B 和 S 边（与 Java 对齐：只保留 C 边成环）
    H_basis = G.copy()
    H_basis.remove_edges_from(skip_edges)
    if len(H_basis) > 0:
        # cycle_basis 返回独立环的节点序列，每个环不闭合，但 cycle[0] 是起点
        basis_cycles = nx.cycle_basis(H_basis)
    else:
        basis_cycles = []

    # 转换为节点序列闭合形式（与 Java recognizeLoopNew 输出格式一致）
    cycles = []
    for cyc in basis_cycles:
        if len(cyc) >= 3:  # 至少 3 个节点成环
            # 闭合：起点加到末尾形成 cycle
            closed = list(cyc) + [cyc[0]]
            cycles.append(closed)

    # ★ 兜底：如果 cycle_basis 没找到，用连通分量数判断
    # 环数 = E - V + C (C=连通分量数)，这是基本环的数学公式
    if not cycles and len(break_edges) > 0:
        H_check = G.copy()
        H_check.remove_edges_from(skip_edges)
        V = H_check.number_of_nodes()
        E = H_check.number_of_edges()
        C = nx.number_connected_components(H_check)
        fundamental_count = E - V + C  # 环基大小
        print(f"  [INFO] cycle_basis 未找到环，但数学上应有 {fundamental_count} 个基本环")

    return G, pos, break_edges, cycles, components


def _in_same_component(G, u, v):
    """u 和 v 是否在 G 的同一连通分量"""
    for comp in nx.connected_components(G):
        if u in comp:
            return v in comp
    return False


def collect_cycle_branch_ids(G, cycles):
    """汇总每个闭环的分支 id（去重保序），返回 list[list[str]]"""
    result = []
    for c in cycles:
        edge_ids = []
        for j in range(len(c) - 1):
            u, v = c[j], c[j + 1]
            if G.has_edge(u, v):
                eid = G[u][v].get('edge_id', '')
                if eid and eid not in edge_ids:
                    edge_ids.append(eid)
        result.append(edge_ids)
    return result


def print_cycle_branch_ids(G, cycles):
    """在控制台清晰输出每个闭环的分支 id"""
    cycle_ids = collect_cycle_branch_ids(G, cycles)
    print("\n" + "=" * 70)
    print(f"[闭环分支 ID 汇总] 共 {len(cycle_ids)} 个闭环")
    print("=" * 70)
    if not cycle_ids:
        print("  (无闭环)")
    for i, ids in enumerate(cycle_ids):
        print(f"  Loop-{i+1:>2}  边数={len(ids):>2}  分支ID: {ids}")
    # 统计去重后的总分支数
    all_ids = sorted({i for ids in cycle_ids for i in ids})
    print(f"\n  涉及分支总数（去重）: {len(all_ids)}")
    print(f"  全部闭环分支 ID: {all_ids}")
    print("=" * 70 + "\n")
    return cycle_ids


def draw(G, pos, break_edges, cycles, save_path='cycle.png'):
    fig, ax = plt.subplots(figsize=(24, 18))

    # 状态统计
    status_count = {'B': 0, 'C': 0, 'S': 0, '其他': 0}
    for _, _, d in G.edges(data=True):
        s = d.get('status', '其他')
        status_count[s if s in status_count else '其他'] += 1

    # 画 C 边（绿色 = 保持连通）
    c_edges = [(u, v) for u, v, d in G.edges(data=True) if d['status'] == 'C']
    nx.draw_networkx_edges(G, pos, edgelist=c_edges, ax=ax,
                           edge_color='green', width=1)

    # 画 S 边（黑色 = 单线/绕线）
    s_edges = [(u, v) for u, v, d in G.edges(data=True) if d['status'] == 'S']
    nx.draw_networkx_edges(G, pos, edgelist=s_edges, ax=ax,
                           edge_color='black', width=1.5)

    # B 边（打断）= 完全不显示

    # 画回路（橙色高亮 = 闭环路径上的 C 边）
    for idx, cycle in enumerate(cycles):
        if len(cycle) < 2:
            continue
        cycle_edges = []
        for i in range(len(cycle) - 1):
            u, v = cycle[i], cycle[i + 1]
            if G.has_edge(u, v):
                cycle_edges.append((u, v))
        if not cycle_edges:
            continue
        nx.draw_networkx_edges(G, pos, edgelist=cycle_edges, ax=ax,
                               edge_color='orange', width=4, alpha=0.9)
        # 标记回路编号
        mid = cycle_edges[len(cycle_edges) // 2]
        if mid[0] in pos:
            x, y = pos[mid[0]]
            ax.annotate(f"Loop-{idx+1}", (x, y), color='orange',
                        fontsize=8, weight='bold')

    # 画点（小灰点，不显示名字）
    nx.draw_networkx_nodes(G, pos, ax=ax, node_size=8, node_color='gray',
                           edgecolors='none')
    # 不画 labels（避免方框遮挡）

    ax.set_title('线束拓扑闭环分析\n绿=C连通 / 黑=S单线 / 橙=闭环路径\n(B 打断边不显示)',
                 fontsize=14, weight='bold')
    ax.set_aspect('equal')
    plt.tight_layout()
    plt.savefig(save_path, dpi=200, bbox_inches='tight')
    print(f"[OK] 已保存: {save_path}")
    print(f"[INFO] 状态统计: B={status_count['B']}, C={status_count['C']}, "
          f"S={status_count['S']}, 其他={status_count['其他']}")
    print(f"[INFO] 画出的边: 绿(C)={len(c_edges)}, 黑(S)={len(s_edges)}, "
          f"暗红(B)={len(break_edges)}, 橙(回路)={sum(len(c) for c in cycles)}")
    print(f"[INFO] 闭环数: {len(cycles)}（基本环，与 Java recognizeLoopNew 等价）")
    for i, c in enumerate(cycles):
        edge_ids = []
        for j in range(len(c) - 1):
            u, v = c[j], c[j + 1]
            if G.has_edge(u, v):
                eid = G[u][v].get('edge_id', '')
                if eid and eid not in edge_ids:
                    edge_ids.append(eid)
        print(f"  Loop-{i+1} (边数={len(c)-1}, 分支id={edge_ids}): "
              f"{' -> '.join(c[:8])}{'...' if len(c) > 8 else ''}")
    plt.show()


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='线束拓扑闭环分析')
    parser.add_argument('txt_path', nargs='?',
                        default=r'F:\office\idearProjects\project20251009\src\main\resources\BS4EM项目json优化设置.txt',
                        help='线束 txt 路径')
    parser.add_argument('scheme_path', nargs='?',
                        # default=r'F:\office\idearProjects\project20251009\src\main\resources\测试新遗传算法2266成本.json',
                        help='方案 json 路径(与 --from-excel 二选一)')
    parser.add_argument('--scheme-idx', type=int, default=0,
                        help='方案号,0-based:json 模式取 list[idx],Excel 模式取方案(idx+1)')
    parser.add_argument('--from-excel', type=str, default=r'F:\office\idearProjects\project20251009\src\main\resources\SchemeChangeTrace_9b40fcbd-47b8-4724-b6ce-7b17e5dd1b7b_20260721_162739.xlsx',
                        help='从方案状态变更追踪 Excel(.xlsx)读方案状态,指定后忽略 scheme_path')
    parser.add_argument('--stage', type=str, default='精确前',
                        choices=['精确前', '精确后', '绕线后'],
                        help='--from-excel 模式下选哪个阶段,默认 绕线后')
    args = parser.parse_args()
    txt_path = args.txt_path
    scheme_path = args.scheme_path

    print(f"[INFO] 读 txt: {txt_path}")
    edges, json_obj = load_edges(txt_path)
    # 顺便提取 normList（与 edges 顺序对应，用于坐标排序）
    norm_list = json_obj.get('normList', [e.get('id') for e in edges])
    print(f"[INFO] edges 数: {len(edges)}, normList 数: {len(norm_list)}")

    if args.from_excel:
        print(f"[INFO] 读 Excel: {args.from_excel} (方案 #{args.scheme_idx + 1}, 阶段={args.stage})")
        statuses = load_scheme_from_excel(
            args.from_excel, args.scheme_idx + 1, args.stage)
        print(f"[INFO] Excel 读到 {len(statuses)} 条状态(行序与 normList 一致)")
    else:
        print(f"[INFO] 读方案: {scheme_path}")
        statuses = load_scheme(scheme_path, idx=args.scheme_idx)
        print(f"[INFO] statuses 数: {len(statuses)}（取方案 #{args.scheme_idx}）")

    if len(edges) != len(statuses):
        print(f"[WARN] 长度不一致！edges={len(edges)} != statuses={len(statuses)}")
        print(f"       按 min 取交集: {min(len(edges), len(statuses))}")

    G, pos, break_edges, cycles, components = build_graph_and_find_cycles(edges, statuses)
    print(f"[INFO] B 打断数: {len(break_edges)}")
    print(f"[INFO] 连通分量数（>1 表示有闭环）: {len(components)}")
    if len(components) > 1:
        for i, comp in enumerate(components):
            print(f"  分量-{i+1}: {len(comp)} 个点")

    # ★ 闭环分支 ID 单独汇总输出
    cycle_ids = print_cycle_branch_ids(G, cycles)
    # 落盘：方便后续脚本读取
    try:
        ids_path = os.path.splitext(os.path.basename(txt_path))[0] + "_cycle_branch_ids.json"
        with open(ids_path, 'w', encoding='utf-8') as f:
            json.dump({
                "loopCount": len(cycle_ids),
                "loops": [{"loopId": i + 1, "branchIds": ids} for i, ids in enumerate(cycle_ids)],
                "allBranchIds": sorted({i for ids in cycle_ids for i in ids}),
            }, f, ensure_ascii=False, indent=2)
        print(f"[OK] 闭环分支 ID 已保存: {ids_path}")
    except Exception as e:
        print(f"[WARN] 保存闭环分支 ID 失败: {e}")

    draw(G, pos, break_edges, cycles)
